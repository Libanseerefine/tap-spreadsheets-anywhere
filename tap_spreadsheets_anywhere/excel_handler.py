import re
import time
import logging
import openpyxl

LOGGER = logging.getLogger(__name__)

def get_header_map(header_row):
    """
    Create a mapping of header names (lowercased) to their indices.
    If a cell does not contain a string, a fallback "ColumnX" is used.
    """
    header_map = {}
    for idx, cell in enumerate(header_row):
        header_val = cell.value if isinstance(cell.value, str) else f"Column{idx + 1}"
        header_map[header_val.lower()] = idx
    return header_map

def get_filter_column_indices(filtered_columns, header_row):
    """
    Given the list of filtered column names and the header row,
    return the list of column indices to check for non-empty values.
    """
    header_map = get_header_map(header_row)
    filter_column_indices = []
    for col in filtered_columns:
        col_lower = col.lower().strip()
        # Case 1: direct match
        if col_lower in header_map:
            filter_column_indices.append(header_map[col_lower])
            continue

        # Case 2: match "ColumnX" pattern
        match = re.match(r"^column(\d+)$", col_lower)
        if match:
            col_index = int(match.group(1)) - 1
            if col_index < len(header_row):
                filter_column_indices.append(col_index)
                continue
            else:
                LOGGER.warning("Filtered column '%s' is out of range.", col)
        LOGGER.warning(
            "Filtered column '%s' not found in header and not recognized as 'ColumnX'. Ignoring.",
            col
        )
    return filter_column_indices

def format_header(header_cell, index, encapsulate_with_brackets):
    """
    Returns a formatted header key for a cell.
    If the header cell is not a string, "ColumnX" is used.
    Depending on the flag, the result may be encapsulated with brackets
    or cleaned up by removing non-word characters.
    """
    if isinstance(header_cell.value, str):
        formatted_key = header_cell.value
    else:
        formatted_key = f"Column{index + 1}"
    
    if encapsulate_with_brackets:
        formatted_key = f"[{formatted_key}]"
    else:
        formatted_key = re.sub(r"[^\w\s]", '', formatted_key)
        formatted_key = re.sub(r"\s+", '_', formatted_key)
        formatted_key = formatted_key.lower()
    return formatted_key

def should_include_column(header_cell, index, included_columns, included_columns_lower):
    """
    Checks whether the current column should be included based on the included_columns list.
    """
    if included_columns:
        if isinstance(header_cell.value, str) and header_cell.value in included_columns:
            return True  # Exact match
        else:
            # Generate a fallback formatted header and test for any substring match
            formatted_key = (header_cell.value.lower() 
                             if isinstance(header_cell.value, str) 
                             else f"Column{index + 1}")
            formatted_key = re.sub(r"\s+", '_', formatted_key)
            return any(inc in formatted_key for inc in included_columns_lower)
    return True

def should_exclude_column(header_cell, index, excluded_columns_lower):
    """
    Checks whether the current column should be excluded based on the excluded_columns set.
    It tests both the formatted header and the "columnX" version.
    """
    formatted_key = (re.sub(r"\s+", '_', header_cell.value.lower())
                     if isinstance(header_cell.value, str)
                     else f"column{index+1}")
    column_index_key = f"column{index+1}"
    if formatted_key in excluded_columns_lower or column_index_key in excluded_columns_lower:
        return True
    return False

def generator_wrapper(reader, encapsulate_with_brackets=False, excluded_columns=None, 
                      skip_initial=0, included_columns=None, filtered_columns=None, rename_mapping=None):
    """
    Processes rows from the reader and yields a dictionary for each row
    after applying filtering and formatting rules.
    Timing is measured on key operations.
    """

    start_total = time.perf_counter()

    # Set default parameters and preformat lists
    filtered_columns = filtered_columns or []
    # If included_columns is specified, ignore excluded_columns
    excluded_columns = [] if excluded_columns is None or included_columns else excluded_columns
    included_columns = included_columns or []
    included_columns_lower = [col.lower() for col in included_columns]
    excluded_columns_lower = {re.sub(r"\s+", '_', col.lower()) for col in excluded_columns}

    _skip_count = 0
    header_row = None

    for row in reader:
        t0 = time.perf_counter()

        # Skip initial rows if needed
        if _skip_count < skip_initial:
            LOGGER.debug("Skipped (%d/%d) row: %r", _skip_count, skip_initial, row)
            _skip_count += 1
            continue
        t1 = time.perf_counter()
        LOGGER.debug("Time to check skip condition: %f seconds", t1 - t0)

        # Identify header row
        if header_row is None:
            header_row = row
            t_header = time.perf_counter()
            LOGGER.debug("Header row set. Time elapsed: %f seconds", t_header - t1)
            continue

        # Check filtered_columns (if any)
        t_filter_start = time.perf_counter()
        if filtered_columns:
            filter_column_indices = get_filter_column_indices(filtered_columns, header_row)
            if all(not row[i].value for i in filter_column_indices):
                LOGGER.debug("Row skipped due to empty values in filtered_columns '%s': %r", filtered_columns, row)
                continue
        t_filter_end = time.perf_counter()
        LOGGER.debug("Time for filtered_columns check: %f seconds", t_filter_end - t_filter_start)

        # Process each cell in the row
        to_return = {}
        for index, cell in enumerate(row):
            t_cell_start = time.perf_counter()

            header_cell = header_row[index]

            # Create a formatted header key
            formatted_key = format_header(header_cell, index, encapsulate_with_brackets)

            # Check included columns (if specified)
            if included_columns:
                if not should_include_column(header_cell, index, included_columns, included_columns_lower):
                    continue
            # Otherwise, check excluded columns if provided
            elif excluded_columns:
                if should_exclude_column(header_cell, index, excluded_columns_lower):
                    continue

            # Final formatting step (reapply cleaning if needed)
            formatted_key = re.sub(r"[^\w\s]", '', formatted_key).replace(' ', '_')

            if rename_mapping:
                key_for_rename = formatted_key.lower()
                if key_for_rename in rename_mapping:
                    formatted_key = rename_mapping[key_for_rename]

            to_return[formatted_key] = cell.value

            t_cell_end = time.perf_counter()
            LOGGER.debug("Time to process cell %d: %f seconds", index, t_cell_end - t_cell_start)

        t_row_end = time.perf_counter()
        LOGGER.debug("Time to process row: %f seconds", t_row_end - t0)
        yield to_return

    end_total = time.perf_counter()
    LOGGER.info("Total processing time: %f seconds", end_total - start_total)

def get_row_iterator(table_spec, file_handle):
    encapsulate_with_brackets = table_spec.get('encapsulate_with_brackets', False)
    excluded_columns = table_spec.get('excluded_columns', [])
    included_columns = table_spec.get('included_columns', [])
    skip_initial = table_spec.get("skip_initial", 0)
    filtered_columns = table_spec.get("filtered_columns", 0)
    rename_mapping = table_spec.get("rename_mapping", {})
    workbook = openpyxl.load_workbook(file_handle.name, read_only=True, data_only = True)
    if "worksheet_name" in table_spec:
        possible_sheet_names = [name.strip() for name in table_spec["worksheet_name"].split(',')]
        active_sheet = None
        
        for sheet_name in possible_sheet_names:
            if sheet_name in workbook.sheetnames:
                # Found a matching worksheet in the workbook
                active_sheet = workbook[sheet_name]
                break

        if not active_sheet:
            # None of the candidate sheets were found
            LOGGER.error(
                "Unable to open any of the specified sheets '%s'. "
                "Did you check for typos or extra spaces?",
                table_spec["worksheet_name"],
            )
            raise ValueError(
                f"No valid sheet found in workbook from list: {table_spec['worksheet_name']}"
            )
    else:
        try:
            worksheets = workbook.worksheets
            if len(worksheets) == 1:
                active_sheet = worksheets[0]
            else:
                max_row = 0
                longest_sheet_index = 0
                for i, sheet in enumerate(worksheets):
                    if sheet.max_row > max_row:
                        max_row = i.max_row
                        longest_sheet_index = i
                active_sheet = worksheets[longest_sheet_index]
        except Exception as e:
            LOGGER.info(e)
            active_sheet = worksheets[0]
    return generator_wrapper(active_sheet, encapsulate_with_brackets, excluded_columns, skip_initial, included_columns, filtered_columns, rename_mapping)